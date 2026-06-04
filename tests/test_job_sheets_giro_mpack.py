"""Focused tests for the GIRO / MPACK / tomorrow-date extensions.

Runnable without pytest:  python tests/test_job_sheets_giro_mpack.py
(also collected by pytest if it is installed).
"""

from __future__ import annotations

import io
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from job_sheets.lines import is_allowed, line_kind, normalize_line  # noqa: E402
from job_sheets.parser import parse_pasted_rows  # noqa: E402
from job_sheets.pd080_generator import generate_xlsx  # noqa: E402
from job_sheets.processor import process_rows  # noqa: E402


def _cells(ws):
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and str(c.value).strip() != "":
                out[c.coordinate] = c.value
    return out


def test_line_mappings():
    cases = {
        "TOP SEAL": "TS", "TP SEAL2": "TS2", "TP SEAL4": "TS4",
        "TP SEAL5": "TS5", "TP SEAL6": "TS6", "TP SEAL7": "TS7",
        "FLOW 5": "F5", "FLOW 7": "F7", "LINERLESS": "LL",
        "GIRO 3": "G3", "GIRO 4": "G4", "GIRO 5": "G5",
        "MPACK 1": "M1", "MPACK 2": "M2", "MPACK 3": "M3", "MPACK 4": "MP4",
    }
    for wc, expected in cases.items():
        assert is_allowed(wc), f"{wc} should be allowed"
        assert normalize_line(wc) == expected, f"{wc} -> {normalize_line(wc)} != {expected}"
    assert line_kind("GIRO 5") == "giro"
    assert line_kind("MPACK 3") == "mpack"
    assert line_kind("TOP SEAL") == "standard"
    assert not is_allowed("H/LINES") and not is_allowed("OFF SITE")
    print("ok test_line_mappings")


def test_topseal_still_works():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Mango 6X2 Perf Ripe\tLoose to Punnet\tTOP SEAL\t1683047\t10",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    assert len(sheets) == 1
    s = sheets[0]
    assert s.kind == "standard"
    assert s.line == "TS"
    assert s.product == "Mango"
    assert s.needs_punnets and s.punnet_material == "2 Cavity"
    rows = s.sheet_rows()
    assert rows[0]["row"] == 5 and rows[0]["product"] == "Mango"
    assert rows[1]["row"] == 12 and rows[1]["product"] == "2 Cavity"
    print("ok test_topseal_still_works")


def test_mpack_two_rows_and_merge():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Oranges Large 8x6 MPACK (dtd)\t\tMPACK 3\t1683001\t100",
        "Oranges Large 8x6 MPACK (dtd)\t\tMPACK 3\t1683002\t50",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    assert len(sheets) == 1, f"duplicate MPACK jobs should merge, got {len(sheets)}"
    s = sheets[0]
    assert s.kind == "mpack"
    assert s.line == "M3"
    assert s.product == "Large Orange"
    assert s.job_number == "1683001"  # largest Act.Qty wins
    assert s.header_suffix.strip() == "/3002"  # suffix top-right, existing behaviour
    rows = s.sheet_rows()
    assert [r["row"] for r in rows] == [5, 6]
    assert all(r["product"] == "Large Orange" for r in rows)
    print("ok test_mpack_two_rows_and_merge")


def test_giro_front_back_net():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Lime - 12X5\t\tGIRO 5\t1683010\t40",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    assert len(sheets) == 1
    s = sheets[0]
    assert s.kind == "giro"
    assert s.line == "G5"
    assert s.product == "Ess Limes"      # Using name from GIRO rules file
    assert s.giro_front == "Ess Limes"   # Front column
    assert s.giro_back == "2 for 3"      # Back column
    assert s.giro_net == "Green Net"     # Net column, display-formatted
    assert s.status == "matched"
    rows = s.sheet_rows()
    assert [r["row"] for r in rows] == [5, 12, 19]
    assert [r["product"] for r in rows] == ["Ess Limes", "2 for 3", "Green Net"]
    print("ok test_giro_front_back_net")


def test_giro_merge_same_product():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Satsuma 14x1Kg\t\tGIRO 4\t1683020\t30",
        "Satsuma 14x1Kg\t\tGIRO 4\t1683021\t10",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    assert len(sheets) == 1, f"duplicate GIRO jobs should merge, got {len(sheets)}"
    s = sheets[0]
    assert s.product == "Satsuma 1kg"
    assert s.giro_front == "Satsuma 1 kg"
    assert s.giro_net == "Orange Net"
    assert s.header_suffix.strip() == "/3021"
    print("ok test_giro_merge_same_product")


def test_generated_date_is_tomorrow_in_xlsx():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Lime - 12X5\t\tGIRO 5\t1683010\t40",
        "Oranges Large 8x6 MPACK (dtd)\t\tMPACK 3\t1683001\t100",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    xlsx = generate_xlsx(sheets)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    tomorrow = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%y")
    for ws in wb.worksheets:
        banner = ws["B1"].value or ""
        assert banner.startswith(tomorrow), f"banner {banner!r} should start with {tomorrow}"
    print(f"ok test_generated_date_is_tomorrow_in_xlsx (date={tomorrow})")


def test_xlsx_layout_matches_examples():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Lime - 12X5\t\tGIRO 5\t\t40",
        "Oranges Large 8x6 MPACK (dtd)\t\tMPACK 3\t\t100",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    by_kind = {s.kind: s for s in sheets}
    xlsx = generate_xlsx(sheets)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    # One sheet per job, landscape, single page.
    assert len(wb.worksheets) == len(sheets)

    # Find each sheet by its line code in B5.
    for ws in wb.worksheets:
        cells = _cells(ws)
        if cells.get("B5") == "G5":
            assert cells.get("C5") == "Ess Limes"
            assert cells.get("B12") == "G5" and cells.get("C12") == "2 for 3"
            assert cells.get("B19") == "G5" and cells.get("C19") == "Green Net"
            assert "C6" not in cells  # GIRO leaves rows 6-11 blank
        elif cells.get("B5") == "M3":
            assert cells.get("C5") == "Large Orange"
            assert cells.get("B6") == "M3" and cells.get("C6") == "Large Orange"
            assert "C12" not in cells  # MPACK only fills rows 5-6
    print("ok test_xlsx_layout_matches_examples")


def test_job_number_font_matches_and_shrinks():
    text = "\n".join([
        "Material Description\tText\tWork Center\tProd. Order\tAct.Qty",
        "Mango 6X2 Perf Ripe\tLoose to Punnet\tTOP SEAL\t1683047\t10",
    ])
    sheets = process_rows(parse_pasted_rows(text))
    xlsx = generate_xlsx(sheets)
    ws = openpyxl.load_workbook(io.BytesIO(xlsx)).active
    main_font = ws["C5"].font       # bold Calibri 14 product entry
    job = ws["E5"]
    assert job.value == "1683047"
    # Same font family/style as the other entries...
    assert job.font.name == main_font.name
    assert job.font.size == main_font.size
    assert job.font.bold == main_font.bold
    # ...but auto-fit so long numbers never overflow / wrap.
    assert job.alignment.shrink_to_fit is True
    assert not job.alignment.wrap_text
    print("ok test_job_number_font_matches_and_shrinks")


def main():
    test_line_mappings()
    test_topseal_still_works()
    test_mpack_two_rows_and_merge()
    test_giro_front_back_net()
    test_giro_merge_same_product()
    test_generated_date_is_tomorrow_in_xlsx()
    test_xlsx_layout_matches_examples()
    test_job_number_font_matches_and_shrinks()
    print("\nALL TESTS PASSED")


if __name__ == "__main__":
    main()
