"""Generate PD080 sheets by filling the REAL template — no redrawing.

Approach (high fidelity, preserves everything):
* The combined workbook is built from ``resources/templates/PD080.xlsx``.
* One worksheet per job is produced by ``copy_worksheet`` of the pristine
  template sheet. ``copy_worksheet`` preserves cell styles, merged cells,
  column widths, row heights, borders, fonts and page setup, but it DROPS the
  logo image and the print area — so we re-attach the logo (extracted from the
  template's ``xl/media``) and re-set the print area on every sheet.
* Only data cells are written; rows 7-20 are never touched, so each sheet keeps
  at least 7 blank usable rows for manual additions.

Output:
* :func:`generate_xlsx` — one print-ready workbook (primary deliverable).
* :func:`generate_pdf`  — one combined PDF, IF a converter (LibreOffice
  ``soffice`` or Excel via win32com) is available; otherwise ``None``.

Cell map (per sheet) — matches resources/example/example.xlsx (sheet 1683047),
the manually-corrected source of truth. Only these cells are written; all others
are left exactly as the template:
    B1 (merged B1:M1)  one left-aligned banner string laid out by space-padding:
        "<date>      <product>      Packaging Sign Out & In Record      <suffix>"
        e.g. "03.06.26           Mango      ...In Record       /3046"
    row 5   main product line  -> B5 Line | C5 Product | D5 Trace | E5 Job Number
    row 12  punnet line (opt.) -> B12 Line | C12 Product | D12 Trace | E12 Job | F12 Qty
        (7-row gap; rows 6-11 stay blank for manual additions)
    Only the punnet/material row gets a Qty, written into F = Sign-out Qty.
    The print date (current date) is written into the banner.
"""

from __future__ import annotations

import copy
import io
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties

from .processor import JobSheet

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent / "resources" / "templates" / "PD080.xlsx"
)

PRINT_AREA = "A1:N24"
PAPER_A4 = 9  # openpyxl paperSize code for A4

MAIN_ROW = 5  # main product row
PUNNET_ROW = MAIN_ROW + 7  # 12 — 7-row gap leaves rows 6-11 blank for additions
QTY_COLUMN = "F"  # Sign-out > Qty (punnet/material row only)

TITLE_TEXT = "Packaging Sign Out & In Record"

# Character columns the banner segments start at, reproducing example.xlsx
# (sheet 1683047) exactly for a short product: date.ljust(19) + product.ljust(35)
# + TITLE.ljust(37) + suffix  ->  the suffix begins at char 91 (== len 96 there).
_BANNER_DATE_W = 19  # product starts at col 19
_BANNER_PRODUCT_W = 35  # title starts at col 19 + 35 = 54
_BANNER_TITLE_W = 37  # suffix starts at col 54 + 37 = 91

_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def _banner(date_str: str, product: str, suffix: str) -> str:
    """Build the single B1 banner string laid out with space-padding.

    ``suffix`` is the additional-job suffix without its leading space, e.g.
    ``"/3046"`` or ``"/3046 /0727"`` (empty when there are no additional jobs).
    """
    text = (
        date_str.ljust(_BANNER_DATE_W)
        + (product or "").ljust(_BANNER_PRODUCT_W)
        + TITLE_TEXT.ljust(_BANNER_TITLE_W)
        + (suffix or "")
    )
    return text.rstrip()


# --------------------------------------------------------------------------- #
# Logo handling
# --------------------------------------------------------------------------- #
def _template_logos() -> list[bytes]:
    """Raw bytes of every image embedded in the template (usually one logo)."""
    out: list[bytes] = []
    with zipfile.ZipFile(TEMPLATE_PATH) as z:
        for name in sorted(z.namelist()):
            if name.startswith("xl/media/"):
                out.append(z.read(name))
    return out


def _safe_title(base: str, used: set[str], index: int) -> str:
    title = _INVALID_SHEET_CHARS.sub("", base or "").strip() or f"Job {index + 1}"
    title = title[:31]
    candidate = title
    n = 2
    while candidate.lower() in used:
        suffix = f" ({n})"
        candidate = title[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate.lower())
    return candidate


def _copy_cell_format(src, dst) -> None:
    """Copy the visible formatting (font + alignment) from one cell to another.

    Used so the punnet row (template default Calibri 11) matches the main row's
    Calibri 14 bold, exactly like the corrected example. Borders/fills are left
    as the template's (the grid is identical on every data row)."""
    dst.font = copy.copy(src.font)
    dst.alignment = copy.copy(src.alignment)


def _style_job_cell(cell, ref) -> None:
    """Style a Job Number cell to match the rest of the sheet, but auto-fit.

    * Font family/style is copied from ``ref`` (the bold Calibri 14 product
      cell) so the Job Number looks visually consistent with the other entries.
    * ``shrink_to_fit`` lets Excel/LibreOffice automatically scale the text down
      when a long Prod.Order would otherwise overflow the cell — it never wraps
      onto a second line and never spills past the border, keeping the printed
      layout clean. Short job numbers stay at the full font size.
    """
    cell.font = copy.copy(ref.font)
    keep = cell.alignment
    cell.alignment = Alignment(
        horizontal=keep.horizontal or "center",
        vertical=keep.vertical or "center",
        shrink_to_fit=True,
        wrap_text=False,
    )


def _apply_page_setup(ws) -> None:
    """Force the sheet to print as exactly ONE landscape A4 page.

    Guarantees one PD080 = one PDF page (and one printed page). fit-to-page only
    shrinks to fit, never enlarges, so the template layout is preserved — it just
    can't spill the print area (A1:N24) onto a second page."""
    ws.print_area = PRINT_AREA
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = PAPER_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    # fitToWidth/Height only take effect when fitToPage is enabled.
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)


def _fill_sheet(ws, sheet: JobSheet, date_str: str) -> None:
    # Top banner: date + product + title + suffix in the single merged B1:M1
    # cell, laid out by space-padding and left-aligned (matches example.xlsx).
    ws["B1"] = _banner(date_str, sheet.product, sheet.header_suffix.strip())
    b1 = ws["B1"]
    b1.alignment = Alignment(horizontal="left", vertical="center")

    # Each row carries its own template row number (set per line family in
    # JobSheet.sheet_rows): standard = rows 5/12, MPACK = rows 5/6,
    # GIRO = rows 5/12/19 (Front/Back/Net).
    for line in sheet.sheet_rows():
        r = line["row"]
        ws[f"B{r}"] = line["line"]
        ws[f"C{r}"] = line["product"]
        ws[f"D{r}"] = line["trace"]
        # Job Number comes ONLY from Prod.Order: leave the cell blank if absent.
        if line["job"]:
            ws[f"E{r}"] = line["job"]
            # Match the bold product font, then shrink-to-fit so long Prod.Orders
            # stay inside the cell without wrapping or overflowing the border.
            _style_job_cell(ws[f"E{r}"], ws[f"C{MAIN_ROW}"])
        # Qty only on the punnet/material row.
        if line["qty"]:
            ws[f"{QTY_COLUMN}{r}"] = (
                int(line["qty"]) if str(line["qty"]).isdigit() else line["qty"]
            )
        # Make every extra row look like the main row (bold product/line/trace).
        if r != MAIN_ROW:
            for col in ("B", "C", "D"):
                _copy_cell_format(ws[f"{col}{MAIN_ROW}"], ws[f"{col}{r}"])


def generate_xlsx(sheets: list[JobSheet]) -> bytes:
    """Build one workbook with a filled copy of the PD080 template per job."""
    base = load_workbook(TEMPLATE_PATH)
    template_ws = base.active

    # Capture original logo anchors, then strip images from the template sheet
    # (copies wouldn't carry them anyway); we re-add fresh ones to every sheet.
    logo_bytes = _template_logos()
    anchors = [copy.deepcopy(im.anchor) for im in template_ws._images]
    template_ws._images = []

    # Make all sheet copies from the PRISTINE template before filling anything.
    targets = [template_ws]
    for _ in range(1, max(1, len(sheets))):
        targets.append(base.copy_worksheet(template_ws))

    # Date written into every banner. Picking is always done for the NEXT day,
    # so the sheet date is ALWAYS tomorrow (current date + 1 day), never today,
    # never a SAP/template date. dd.mm.yy as in the example golden files.
    date_str = (datetime.now() + timedelta(days=1)).strftime("%d.%m.%y")

    # Persist logos to temp files so openpyxl can read them reliably at save.
    tmp_files: list[str] = []
    try:
        logo_paths: list[str] = []
        for data in logo_bytes:
            fd, path = tempfile.mkstemp(suffix=".img")
            os.close(fd)
            with open(path, "wb") as f:
                f.write(data)
            tmp_files.append(path)
            logo_paths.append(path)

        used_titles: set[str] = set()
        for i, sheet in enumerate(sheets):
            ws = targets[i]
            ws.title = _safe_title(sheet.job_number or sheet.product, used_titles, i)
            _apply_page_setup(ws)
            for path, anchor in zip(logo_paths, anchors):
                img = XLImage(path)
                img.anchor = copy.deepcopy(anchor)
                ws.add_image(img)
            _fill_sheet(ws, sheet, date_str)

        # If there were zero sheets we still saved the bare template; guard it.
        if not sheets:
            _apply_page_setup(template_ws)

        buffer = io.BytesIO()
        base.save(buffer)
        buffer.seek(0)
        return buffer.read()
    finally:
        for path in tmp_files:
            try:
                os.remove(path)
            except OSError:
                pass


# --------------------------------------------------------------------------- #
# Optional XLSX -> PDF conversion
# --------------------------------------------------------------------------- #
def _registry_soffice() -> list[str]:
    """LibreOffice path(s) from the Windows registry App Paths key."""
    try:
        import winreg
    except Exception:
        return []
    out: list[str] = []
    key = r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\soffice.exe"
    roots = (winreg.HKEY_LOCAL_MACHINE, winreg.HKEY_CURRENT_USER)
    views = (getattr(winreg, "KEY_WOW64_64KEY", 0), getattr(winreg, "KEY_WOW64_32KEY", 0))
    for root in roots:
        for view in views:
            try:
                with winreg.OpenKey(root, key, 0, winreg.KEY_READ | view) as k:
                    val, _ = winreg.QueryValueEx(k, None)  # default value = full path
                    if val:
                        out.append(val)
            except OSError:
                continue
    return out


def _soffice_candidates() -> list[str]:
    """All places LibreOffice's soffice may live, in priority order."""
    program_files = os.environ.get("ProgramFiles", r"C:\Program Files")
    program_files_x86 = os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")
    local = os.environ.get("LOCALAPPDATA", "")
    candidates = [
        # 1) Explicit override for non-standard installs.
        os.environ.get("SOFFICE_PATH", ""),
        os.environ.get("LIBREOFFICE_PATH", ""),
        # 2) On PATH.
        shutil.which("soffice") or "",
        shutil.which("soffice.exe") or "",
        shutil.which("libreoffice") or "",
        # 3) Standard Windows install locations (system-wide + per-user).
        os.path.join(program_files, "LibreOffice", "program", "soffice.exe"),
        os.path.join(program_files_x86, "LibreOffice", "program", "soffice.exe"),
        os.path.join(local, "Programs", "LibreOffice", "program", "soffice.exe") if local else "",
        # 4) Linux/macOS fallbacks (harmless on Windows).
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    ]
    candidates += _registry_soffice()
    return [c for c in candidates if c]


def _find_soffice() -> str | None:
    """Locate the LibreOffice ``soffice`` executable, or None if not installed."""
    for path in _soffice_candidates():
        if os.path.exists(path):
            return path
    return None


def _convert_with_soffice(xlsx_bytes: bytes) -> bytes | None:
    """Convert the combined workbook to ONE multi-page PDF via LibreOffice.

    Each sheet's print area + page setup (landscape A4, fit-to-one-page) is
    honoured by the Calc PDF export filter, so one sheet = one page, in order.

    Uses a throwaway user profile (``-env:UserInstallation``) so conversion
    works even when the operator already has LibreOffice open — a common cause
    of silent headless failures on Windows.
    """
    soffice = _find_soffice()
    if not soffice:
        return None
    workdir = tempfile.mkdtemp(prefix="pd080_")
    try:
        src = os.path.join(workdir, "sheets.xlsx")
        with open(src, "wb") as f:
            f.write(xlsx_bytes)

        profile = os.path.join(workdir, "profile")
        user_installation = "-env:UserInstallation=file:///" + profile.replace("\\", "/")
        result = subprocess.run(
            [
                soffice,
                user_installation,
                "--headless",
                "--norestore",
                "--nolockcheck",
                "--nodefault",
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                workdir,
                src,
            ],
            timeout=180,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )
        pdf_path = os.path.join(workdir, "sheets.pdf")
        if result.returncode == 0 and os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                return f.read()
        return None
    except Exception:
        return None
    finally:
        shutil.rmtree(workdir, ignore_errors=True)


def _convert_with_excel(xlsx_bytes: bytes) -> bytes | None:
    """Windows fallback: drive Excel via COM if installed (pywin32)."""
    try:
        import pythoncom  # noqa: F401
        import win32com.client  # type: ignore
    except Exception:
        return None

    workdir = tempfile.mkdtemp(prefix="pd080_")
    excel = None
    wb = None
    try:
        src = os.path.join(workdir, "sheets.xlsx")
        dst = os.path.join(workdir, "sheets.pdf")
        with open(src, "wb") as f:
            f.write(xlsx_bytes)
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(src)
        wb.ExportAsFixedFormat(0, dst)  # 0 = xlTypePDF
        if os.path.exists(dst):
            with open(dst, "rb") as f:
                return f.read()
        return None
    except Exception:
        return None
    finally:
        try:
            if wb is not None:
                wb.Close(False)
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        shutil.rmtree(workdir, ignore_errors=True)


def pdf_available() -> bool:
    """True if a converter (LibreOffice or Excel) is present for combined PDF."""
    if _find_soffice() is not None:
        return True
    try:  # Windows: Excel via COM (pywin32).
        import win32com.client  # type: ignore  # noqa: F401

        return True
    except Exception:
        return False


def generate_pdf(sheets: list[JobSheet]) -> bytes | None:
    """Combined PDF of all sheets, or None if no converter is available.

    Converts the filled template workbook (so the PDF looks exactly like the
    template). Tries LibreOffice first, then Excel COM.
    """
    xlsx_bytes = generate_xlsx(sheets)
    return _convert_with_soffice(xlsx_bytes) or _convert_with_excel(xlsx_bytes)
