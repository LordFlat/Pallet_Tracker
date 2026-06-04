"""GIRO netting rules — the source of truth for GIRO Front / Back / Net values.

Rules file: resources/rules/Giro.xlsx
Columns: "Material Description", "Using name", "Front", "Back", "Net".

A GIRO PD080 fills exactly three rows from one matched rule:
    Front  (top data row)      -> the netted product label
    Back   (middle data row)   -> the offer / multibuy text (e.g. "2 for 3")
    Net    (penultimate row)   -> the net colour, displayed as "<colour> Net"

Front/Back/Net must come from this file — never guessed from the SAP product
description. Matching reuses the same normalize + fuzzy scoring as the main
material rules so OCR noise and pack-size drift are tolerated.
"""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from .matcher import MATCH_THRESHOLD, _score
from .text_utils import normalize

RULES_PATH = (
    Path(__file__).resolve().parent.parent / "resources" / "rules" / "Giro.xlsx"
)

# Front/Back/Net may be intentionally blank. Operators and spreadsheet exports
# sometimes type a placeholder to MEAN blank; treat those as empty so they never
# reach the PD080 template as literal text.
_BLANK_TOKENS = {"none", "nan", "n/a", "-", "–", "—"}


def _blank_if_placeholder(value: str) -> str:
    return "" if value.strip().lower() in _BLANK_TOKENS else value


@dataclass
class GiroRule:
    material_description: str
    using_name: str
    front: str
    back: str
    net: str
    normalized: str


@dataclass
class GiroMatch:
    rule: GiroRule | None
    score: float

    @property
    def matched(self) -> bool:
        return self.rule is not None and self.score >= MATCH_THRESHOLD


def _header_index(header: tuple) -> dict:
    """Map columns by (loosely matched) header name so column order can drift."""
    idx: dict[str, int] = {}
    for i, name in enumerate(header):
        key = (str(name or "")).strip().lower()
        if key.startswith("material"):
            idx["desc"] = i
        elif key.startswith("using"):
            idx["using"] = i
        elif key.startswith("front"):
            idx["front"] = i
        elif key.startswith("back"):
            idx["back"] = i
        elif key.startswith("net"):
            idx["net"] = i
    return idx


@lru_cache(maxsize=1)
def load_rules() -> tuple[GiroRule, ...]:
    """Load and cache the GIRO rules. Returns an empty tuple if missing/empty.

    The GIRO workbook can carry several (sometimes blank) sheets, so we scan
    for the first one whose header row exposes the expected columns.
    """
    if not RULES_PATH.exists():
        return ()

    wb = load_workbook(RULES_PATH, read_only=True, data_only=True)
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            continue
        idx = _header_index(header)
        if "desc" not in idx:
            continue

        def cell(row, key):
            i = idx.get(key)
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        rules: list[GiroRule] = []
        for row in rows:
            if not row:
                continue
            desc = cell(row, "desc")
            if not desc:
                continue
            rules.append(
                GiroRule(
                    material_description=desc,
                    using_name=cell(row, "using"),
                    front=_blank_if_placeholder(cell(row, "front")),
                    back=_blank_if_placeholder(cell(row, "back")),
                    net=_blank_if_placeholder(cell(row, "net")),
                    normalized=normalize(desc),
                )
            )
        if rules:
            return tuple(rules)
    return ()


def match_giro(raw_description: str) -> GiroMatch:
    """Fuzzy-match a raw SAP description against the GIRO rules.

    Returns the best candidate plus its score. Callers should check
    ``GiroMatch.matched`` before trusting it; an uncertain match must be sent
    to manual review, never guessed silently.
    """
    rules = load_rules()
    target = normalize(raw_description)
    if not target or not rules:
        return GiroMatch(rule=None, score=0.0)

    for rule in rules:  # exact normalized match wins outright
        if rule.normalized == target:
            return GiroMatch(rule=rule, score=100.0)

    best: GiroRule | None = None
    best_score = -1.0
    best_len_gap: int | None = None
    for rule in rules:
        s = _score(target, rule.normalized)
        len_gap = abs(len(rule.normalized) - len(target))
        if s > best_score or (
            s == best_score and (best_len_gap is None or len_gap < best_len_gap)
        ):
            best_score = s
            best_len_gap = len_gap
            best = rule

    return GiroMatch(rule=best, score=max(best_score, 0.0))
