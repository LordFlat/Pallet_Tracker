"""Fuzzy matching of SAP Material Descriptions against the rules database.

Rules file: resources/rules/material_description_task.xlsx
Columns: "Material Description", "Using Name", "Punnets".
"""

from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from .text_utils import normalize

# Score (0-100) at/above which a fuzzy match is treated as confident.
MATCH_THRESHOLD = 78

RULES_PATH = (
    Path(__file__).resolve().parent.parent
    / "resources"
    / "rules"
    / "material_description_task.xlsx"
)

# Prefer rapidfuzz; fall back to the stdlib so the module still works if the
# dependency is missing.
try:  # pragma: no cover - exercised indirectly
    from rapidfuzz import fuzz

    def _score(a: str, b: str) -> float:
        # token_sort_ratio handles word reordering / missing spaces well, while
        # keeping the pack-size digits in play.
        return max(fuzz.token_sort_ratio(a, b), fuzz.partial_ratio(a, b))

except ImportError:  # pragma: no cover
    from difflib import SequenceMatcher

    def _score(a: str, b: str) -> float:
        return SequenceMatcher(None, a, b).ratio() * 100


@dataclass
class Rule:
    material_description: str
    using_name: str
    punnets: str | None
    normalized: str


@dataclass
class MatchResult:
    rule: Rule | None
    score: float

    @property
    def matched(self) -> bool:
        return self.rule is not None and self.score >= MATCH_THRESHOLD


@lru_cache(maxsize=1)
def load_rules() -> tuple[Rule, ...]:
    """Load and cache the rules database. Returns an empty tuple if missing."""
    if not RULES_PATH.exists():
        return ()

    wb = load_workbook(RULES_PATH, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return ()

    # Map columns by (loosely matched) header name so column order can drift.
    idx = {}
    for i, name in enumerate(header):
        key = (str(name or "")).strip().lower()
        if key.startswith("material"):
            idx["desc"] = i
        elif key.startswith("using"):
            idx["using"] = i
        elif key.startswith("punnet"):
            idx["punnets"] = i

    rules: list[Rule] = []
    for row in rows:
        if not row:
            continue
        desc = (str(row[idx["desc"]]) if "desc" in idx and row[idx["desc"]] is not None else "").strip()
        using = (str(row[idx["using"]]) if "using" in idx and row[idx["using"]] is not None else "").strip()
        punnets = None
        if "punnets" in idx and row[idx["punnets"]] is not None:
            punnets = str(row[idx["punnets"]]).strip() or None
        if not desc:
            continue
        rules.append(
            Rule(
                material_description=desc,
                using_name=using,
                punnets=punnets,
                normalized=normalize(desc),
            )
        )

    return tuple(rules)


def match_description(raw_description: str) -> MatchResult:
    """Fuzzy-match a raw SAP description against the rules database.

    Returns the best candidate plus its score. Callers should check
    ``MatchResult.matched`` (score >= MATCH_THRESHOLD) before trusting it; an
    uncertain match must be sent to manual review, never guessed silently.
    """
    rules = load_rules()
    target = normalize(raw_description)
    if not target or not rules:
        return MatchResult(rule=None, score=0.0)

    # Priority 1: an exact normalized match wins outright. This stops a generic
    # rule from tying via partial_ratio — e.g. "PLUM RED 10X400g H/S" must pick
    # "Ess Plum 400" (exact) rather than "MT Plum 340" (whose "10x4" is a partial
    # substring of "10x400g" and would otherwise score 100 too).
    for rule in rules:
        if rule.normalized == target:
            return MatchResult(rule=rule, score=100.0)

    # Priority 2: best fuzzy score. On ties, prefer the candidate whose
    # normalized length is closest to the target (the more specific rule),
    # which also favours the longer pack-size match over a shorter prefix.
    best: Rule | None = None
    best_score = -1.0
    best_len_gap = None
    for rule in rules:
        s = _score(target, rule.normalized)
        len_gap = abs(len(rule.normalized) - len(target))
        if s > best_score or (s == best_score and (best_len_gap is None or len_gap < best_len_gap)):
            best_score = s
            best_len_gap = len_gap
            best = rule

    return MatchResult(rule=best, score=max(best_score, 0.0))
